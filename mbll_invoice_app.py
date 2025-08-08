import streamlit as st
import pandas as pd
import os
import zipfile
import io
import re
from openpyxl import load_workbook

# --- Helper function to sanitize filenames ---
def sanitize_filename(name):
    return re.sub(r'[\\/:"*?<>|]+', '_', str(name))

# --- Invoice Generation Function ---
def process_invoices(order_df, invoice_template):
    # Create pivot table
    pivot_df = order_df.pivot_table(
        index=['Store Name', 'Supplier', 'Supplier Reference', 'PO Number', 'TechPOS Sku'],
        values=['Total Units', 'Unit Cost', 'Total ($)'],
        aggfunc='sum'
    ).reset_index()

    grouped = pivot_df.groupby(['Store Name', 'Supplier Reference'])

    zip_buffer = io.BytesIO()
    summary_data = []

    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        for (store, supplier_ref), group_data in grouped:
            clean_store = sanitize_filename(store)
            clean_supplier_ref = sanitize_filename(supplier_ref)
            folder_name = f"Invoices/{clean_store}"
            file_name = f"{clean_store} - {clean_supplier_ref}.xlsx"
            file_path = f"{folder_name}/{file_name}"

            # Load invoice template
            wb = load_workbook(invoice_template)
            ws = wb.active

            # Write Supplier Reference to B8
            ws['B8'] = supplier_ref

            # Write data from A14
            start_row = 14
            for _, row in group_data.iterrows():
                ws.cell(row=start_row, column=1, value=row['TechPOS Sku'])
                ws.cell(row=start_row, column=2, value=row['Total Units'])
                ws.cell(row=start_row, column=3, value=row['Unit Cost'])
                start_row += 1

            # Save to in-memory buffer
            invoice_buffer = io.BytesIO()
            wb.save(invoice_buffer)
            invoice_buffer.seek(0)

            # Write to zip
            zipf.writestr(file_path, invoice_buffer.read())

            # Add to summary
            summary_data.append({
                'Store Name': store,
                'Supplier Reference': supplier_ref,
                'Total SKUs': group_data['TechPOS Sku'].nunique(),
                'Total Quantity': group_data['Total Units'].sum(),
                'Total Cost': group_data['Total ($)'].sum()
            })

        # Create summary DataFrame and add to zip
        summary_df = pd.DataFrame(summary_data)
        summary_buffer = io.BytesIO()
        summary_df.to_excel(summary_buffer, index=False)
        summary_buffer.seek(0)
        zipf.writestr("MBLL_Invoice_Summary.xlsx", summary_buffer.read())

    zip_buffer.seek(0)
    return summary_df, zip_buffer

# --- Streamlit UI ---
st.set_page_config(page_title="MBLL Invoice Generator", layout="centered")
st.title("📦 MBLL Invoice Generator")

with st.expander("📘 Instructions"):
    st.markdown("""
    **How to use:**
    1. Upload the **MBLL Order Summary.xlsx** file  
    2. Upload the **Invoice Template.xlsx** file  
    3. Click **Generate Invoices**
    4. Download the generated **Summary Excel** and **ZIP of invoices**
    """)

# File uploaders
order_file = st.file_uploader("📁 Upload MBLL Order Summary Excel", type=['xlsx'])
template_file = st.file_uploader("📄 Upload Invoice Template Excel", type=['xlsx'])

# Button to trigger processing
if st.button("🚀 Generate Invoices"):
    if order_file is None or template_file is None:
        st.error("❗ Please upload both files to continue.")
    else:
        with st.spinner("Processing invoices..."):
            order_df = pd.read_excel(order_file)
            summary_df, zip_buffer = process_invoices(order_df, template_file)

        st.success("✅ Invoices and summary generated!")

        # Display summary preview
        st.subheader("📋 Invoice Summary")
        st.dataframe(summary_df)

        # Download buttons
        st.download_button(
            label="📥 Download Summary Excel",
            data=summary_df.to_excel(index=False),
            file_name="MBLL_Invoice_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            label="📦 Download ZIP of Invoices",
            data=zip_buffer,
            file_name="MBLL_Invoices.zip",
            mime="application/zip"
        )
